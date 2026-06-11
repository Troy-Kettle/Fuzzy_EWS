"""Build a per-vital NEWS-2 vs Snapshot Fuzzy Score reference spreadsheet.

For every input value on each vital's membership-function grid, compute:
    - NEWS-2 per-vital score (integer 0-3, from hard thresholds)
    - Snapshot fuzzy score "SS" (continuous 0-3, centroid defuzzification
      over the vital's fuzzy membership functions mapped to concern levels)

Writes a styled workbook with one sheet per vital.

Blood pressure uses the asymmetric (Tab 4) variant: above-normal mild and
moderate concern sets are merged into a single "elevated" → Mild concern
mapping, so Moderate concern is never fired from elevated BP.  Only a
hypertensive crisis (≥220 mmHg) reaches Severe concern from the upper side.

Supplementary oxygen is expressed as nasal-cannula flow rate (L/min) and uses
membership functions derived directly from Part 1 training-data event rates —
no FiO₂ conversion formula is applied.  Gaussian CDF transitions are placed at
empirically justified breakpoints (No concern→Mild at 2 L/min, Mild→Moderate
at 5 L/min, Moderate→Severe at 9 L/min).
NEWS-2 adds +2 whenever any supplemental oxygen is in use
(0 L/min → 0 points; any flow > 0 → +2 points).

Run:  python3 build_vital_score_spreadsheet.py
"""
from __future__ import annotations

import math
import pandas as pd
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from streamlit_app import (
    DATA_DIR_SIGMOID,
    custom_mf_3_var_down,
    custom_mf_3_var_up,
    custom_mf_7_var,
    custom_mf_sbp_sharper,
    defuzz_vital_centroid,
    map_to_concern_levels,
    map_to_concern_levels_sharper_sbp,
)

# ---------------------------------------------------------------------------
# Supplementary oxygen (L/min) membership function generation
# ---------------------------------------------------------------------------
# Membership functions are constructed via Gaussian CDF transitions placed at
# clinically and empirically justified breakpoints derived from Part 1 training
# data event rates (REVIEW_WITHIN_4HOURS).  No L/min → FiO₂ formula is used.
#
# Empirical event rates from Part 1 data:
#   0 L/min (room air): 0.0062   1 L/min: 0.0113   2 L/min: 0.0127
#   3 L/min: 0.0149              4 L/min: 0.0193    5 L/min: 0.0145
#   8 L/min: 0.0280             10 L/min: 0.0303   15 L/min: 0.0513
#
# Transition centres and Gaussian spreads (σ):
#   No concern → Mild     : centre 2 L/min, σ = 1.5
#   Mild → Moderate       : centre 5 L/min, σ = 2.0
#   Moderate → Severe     : centre 9 L/min, σ = 3.0
#
# These parameters mirror the same Gaussian CDF construction used for the FiO₂
# membership functions (verified against the sigmoid CSV), giving a consistent
# methodology while using the L/min scale directly.
#
# Partition of unity is preserved exactly:
#   nc(x) = 1 − Φ((x−t1)/σ1)
#   mild(x) = Φ((x−t1)/σ1) − Φ((x−t2)/σ2)
#   mod(x)  = Φ((x−t2)/σ2) − Φ((x−t3)/σ3)
#   sev(x)  = Φ((x−t3)/σ3)
#   ⇒ nc + mild + mod + sev = 1 for all x

LMIN_MF_FILENAME = "supplementary_oxygen_lmin_membership_functions.csv"

_LMIN_T1, _LMIN_S1 = 2.0, 1.5   # No concern → Mild
_LMIN_T2, _LMIN_S2 = 5.0, 2.0   # Mild → Moderate
_LMIN_T3, _LMIN_S3 = 9.0, 3.0   # Moderate → Severe


def _ncdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2)))


def generate_lmin_mf_csv(output_path: Path) -> None:
    """Generate the supplementary oxygen L/min membership function CSV."""
    rows = []
    for lmin in range(0, 16):
        nc   = 1.0 - _ncdf((lmin - _LMIN_T1) / _LMIN_S1)
        mild = _ncdf((lmin - _LMIN_T1) / _LMIN_S1) - _ncdf((lmin - _LMIN_T2) / _LMIN_S2)
        mod  = _ncdf((lmin - _LMIN_T2) / _LMIN_S2) - _ncdf((lmin - _LMIN_T3) / _LMIN_S3)
        sev  = _ncdf((lmin - _LMIN_T3) / _LMIN_S3)
        rows.append({
            "Value": float(lmin),
            "No concern": nc,
            "Above normal - mild concern": mild,
            "Above normal - moderate concern": mod,
            "Above normal - severe concern": sev,
        })
    pd.DataFrame(rows).to_csv(output_path, index=False)

OUTPUT_PATH = Path(__file__).parent / "vital_score_reference.xlsx"


# ---------------------------------------------------------------------------
# Per-vital NEWS-2 thresholds (Scale 1).
# Duplicated here, rather than reused from streamlit_app.calculate_news2,
# because those scorers are nested inside that function and expect a full
# Observation. Keeping them explicit and per-vital makes the spreadsheet
# self-documenting: the thresholds are visible next to the code that writes them.
# ---------------------------------------------------------------------------

def news_resp(x: float) -> int:
    if x <= 8:
        return 3
    if 9 <= x <= 11:
        return 1
    if 12 <= x <= 20:
        return 0
    if 21 <= x <= 24:
        return 2
    return 3


def news_spo2(x: float) -> int:
    if x <= 91:
        return 3
    if 92 <= x <= 93:
        return 2
    if 94 <= x <= 95:
        return 1
    return 0


def news_temp(x: float) -> int:
    if x <= 35.0:
        return 3
    if 35.1 <= x <= 36.0:
        return 1
    if 36.1 <= x <= 38.0:
        return 0
    if 38.1 <= x <= 39.0:
        return 1
    return 2


def news_bp(x: float) -> int:
    if x <= 90:
        return 3
    if 91 <= x <= 100:
        return 2
    if 101 <= x <= 110:
        return 1
    if 111 <= x <= 219:
        return 0
    return 3


def news_hr(x: float) -> int:
    if x <= 40:
        return 3
    if 41 <= x <= 50:
        return 1
    if 51 <= x <= 90:
        return 0
    if 91 <= x <= 110:
        return 1
    if 111 <= x <= 130:
        return 2
    return 3


def news_fio2_flag(x: float) -> int:
    """NEWS-2 has no per-value FiO2 score; only a +2 supplemental-O2 flag."""
    return 2 if x > 21 else 0


def news_lmin(x: float) -> int:
    """NEWS-2 supplemental O2 flag expressed in L/min: 0 at room air, +2 on any flow."""
    return 2 if x > 0 else 0


# ---------------------------------------------------------------------------
# Vital specifications
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class VitalSpec:
    sheet_name: str
    unit: str
    csv_filename: str
    mf_class: type
    news_fn: Callable[[float], int]
    news_column_header: str
    is_decimal: bool  # True for temperature (one decimal), False otherwise
    news_notes: str
    concern_fn: Optional[Callable] = field(default=None)  # None → map_to_concern_levels


VITALS: list[VitalSpec] = [
    VitalSpec(
        sheet_name="Heart Rate",
        unit="bpm",
        csv_filename="heart_rate_membership_functions.csv",
        mf_class=custom_mf_7_var,
        news_fn=news_hr,
        news_column_header="NEWS-2",
        is_decimal=False,
        news_notes=(
            "NEWS-2 Heart Rate (Scale 1): "
            "≤40→3, 41-50→1, 51-90→0, 91-110→1, 111-130→2, ≥131→3."
        ),
    ),
    VitalSpec(
        sheet_name="Blood Pressure",
        unit="mmHg",
        csv_filename="systolic_blood_pressure_membership_functions.csv",
        mf_class=custom_mf_sbp_sharper,
        news_fn=news_bp,
        news_column_header="NEWS-2",
        is_decimal=False,
        news_notes=(
            "Systolic BP — asymmetric (Tab 4) fuzzy scoring. "
            "Hypotensive side unchanged: ≤90→3, 91-100→2, 101-110→1. "
            "Above normal: mild-raised and moderately-raised fuzzy sets are merged "
            "into a single 'elevated' set mapped to Mild concern only — Moderate "
            "concern is never fired from the upper side. "
            "Only a hypertensive crisis (≥220 mmHg) reaches Severe concern. "
            "NEWS-2 column (unchanged): ≤90→3, 91-100→2, 101-110→1, 111-219→0, ≥220→3."
        ),
        concern_fn=map_to_concern_levels_sharper_sbp,
    ),
    VitalSpec(
        sheet_name="Temperature",
        unit="°C",
        csv_filename="temperature_membership_functions.csv",
        mf_class=custom_mf_7_var,
        news_fn=news_temp,
        news_column_header="NEWS-2",
        is_decimal=True,
        news_notes=(
            "NEWS-2 Temperature (Scale 1): "
            "≤35.0→3, 35.1-36.0→1, 36.1-38.0→0, 38.1-39.0→1, ≥39.1→2."
        ),
    ),
    VitalSpec(
        sheet_name="Respiratory Rate",
        unit="breaths/min",
        csv_filename="respiratory_rate_membership_functions.csv",
        mf_class=custom_mf_7_var,
        news_fn=news_resp,
        news_column_header="NEWS-2",
        is_decimal=False,
        news_notes=(
            "NEWS-2 Respiratory Rate (Scale 1): "
            "≤8→3, 9-11→1, 12-20→0, 21-24→2, ≥25→3."
        ),
    ),
    VitalSpec(
        sheet_name="Oxygen Saturation",
        unit="%",
        csv_filename="oxygen_saturation_membership_functions.csv",
        mf_class=custom_mf_3_var_down,
        news_fn=news_spo2,
        news_column_header="NEWS-2",
        is_decimal=False,
        news_notes=(
            "NEWS-2 SpO2 (Scale 1): ≤91→3, 92-93→2, 94-95→1, ≥96→0."
        ),
    ),
    VitalSpec(
        sheet_name="Supplementary Oxygen",
        unit="L/min",
        csv_filename=LMIN_MF_FILENAME,
        mf_class=custom_mf_3_var_up,
        news_fn=news_lmin,
        news_column_header="NEWS-2 (supp. O2 flag)",
        is_decimal=False,
        news_notes=(
            "Supplementary oxygen flow rate (nasal cannula, L/min). "
            "Membership functions derived directly from Part 1 training-data event rates "
            "using Gaussian CDF transitions (no FiO₂ conversion formula). "
            "Transition centres: No concern→Mild at 2 L/min, Mild→Moderate at 5 L/min, "
            "Moderate→Severe at 9 L/min. "
            "NEWS-2: 0 L/min (room air) → 0 points; any supplemental flow > 0 → +2 points."
        ),
    ),
    VitalSpec(
        sheet_name="Inspired Oxygen (FiO2)",
        unit="% FiO2",
        csv_filename="inspired_oxygen_concentration_membership_functions.csv",
        mf_class=custom_mf_3_var_up,
        news_fn=news_fio2_flag,
        news_column_header="NEWS-2 (supp. O2 flag)",
        is_decimal=False,
        news_notes=(
            "Inspired oxygen concentration (% FiO2) directly from the generated "
            "membership functions (21–100%). "
            "NEWS-2 has no per-value FiO2 score; it adds +2 whenever the patient "
            "is on supplemental oxygen (FiO2 > 21%). This column shows that flag: "
            "0 at 21%, 2 above."
        ),
    ),
]


# ---------------------------------------------------------------------------
# Scoring a single value on one vital
# ---------------------------------------------------------------------------

def compute_row(
    mf,
    value: float,
    news_fn: Callable[[float], int],
    concern_fn: Optional[Callable] = None,
) -> dict:
    """Return one row: input, NEWS score, fuzzy SS, dominant set + strength."""
    _concern = concern_fn if concern_fn is not None else map_to_concern_levels
    memberships = mf(value)
    ss = defuzz_vital_centroid(_concern(memberships))
    dom_label, dom_strength = max(memberships.items(), key=lambda kv: kv[1])
    return {
        "value": float(value),
        "news": int(news_fn(value)),
        "ss": float(ss),
        "dom_label": dom_label,
        "dom_strength": float(dom_strength),
    }


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")  # deep navy
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3A5F")
CAPTION_FONT = Font(name="Calibri", size=10, italic=True, color="444444")
ZEBRA_FILL = PatternFill("solid", fgColor="F4F6FA")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="C9CED6")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(wb: Workbook, spec: VitalSpec) -> None:
    mf = spec.mf_class(DATA_DIR_SIGMOID / spec.csv_filename)
    values = list(mf.df["Value"].values)
    rows = [compute_row(mf, v, spec.news_fn, concern_fn=spec.concern_fn) for v in values]

    ws = wb.create_sheet(title=spec.sheet_name)

    # ----- Title + caption block -----
    n_cols = 5
    last_col = get_column_letter(n_cols)
    ws["A1"] = f"{spec.sheet_name} — NEWS-2 vs Snapshot Fuzzy Score (SS)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{last_col}1")

    ws["A2"] = spec.news_notes
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.merge_cells(f"A2:{last_col}2")
    ws.row_dimensions[2].height = 30

    ws["A3"] = (
        "SS = fuzzy per-vital concern score (0–3), centroid defuzzification of "
        "concern-level firing strengths. The dominant set and its strength "
        "show which fuzzy category drives the input at that value."
    )
    ws["A3"].font = CAPTION_FONT
    ws["A3"].alignment = LEFT
    ws.merge_cells(f"A3:{last_col}3")
    ws.row_dimensions[3].height = 30

    header_row = 5
    headers = [
        f"Input ({spec.unit})",
        spec.news_column_header,
        "Snapshot Score (0–3)",
        "Dominant concern set",
        "Dominant strength (0–1)",
    ]
    col_widths = [14, 10, 22, 36, 22]

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[header_row].height = 28

    # ----- Data rows -----
    value_fmt = "0.0" if spec.is_decimal else "0"
    n_data_cols = 5
    for i, r in enumerate(rows):
        excel_row = header_row + 1 + i
        stripe = (i % 2 == 1)

        ws.cell(row=excel_row, column=1, value=r["value"]).number_format = value_fmt
        ws.cell(row=excel_row, column=2, value=r["news"]).number_format = "0"
        ws.cell(row=excel_row, column=3, value=r["ss"]).number_format = "0.000"
        ws.cell(row=excel_row, column=4, value=r["dom_label"])
        ws.cell(row=excel_row, column=5, value=r["dom_strength"]).number_format = "0.000"

        for col_idx in range(1, n_data_cols + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = CELL_BORDER
            cell.alignment = CENTER if col_idx != 4 else LEFT
            if stripe:
                cell.fill = ZEBRA_FILL

    # ----- Column widths -----
    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_overview_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Overview", index=0)
    ws["A1"] = "Fuzzy EWS — Per-Vital Score Reference"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F3A5F")
    ws.merge_cells("A1:B1")

    ws["A3"] = (
        "This workbook lists, for every input value on each vital's "
        "membership-function grid, the NEWS-2 score and the Snapshot fuzzy "
        "Score (SS) that the model assigns."
    )
    ws["A3"].font = CAPTION_FONT
    ws["A3"].alignment = LEFT
    ws.merge_cells("A3:B3")
    ws.row_dimensions[3].height = 60

    ws["A5"] = "Sheet"
    ws["B5"] = "What it covers"
    for c in ("A5", "B5"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
        ws[c].alignment = CENTER
        ws[c].border = CELL_BORDER
    ws.row_dimensions[5].height = 22

    for i, spec in enumerate(VITALS):
        row = 6 + i
        ws.cell(row=row, column=1, value=spec.sheet_name).border = CELL_BORDER
        ws.cell(row=row, column=2, value=spec.news_notes).border = CELL_BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2).alignment = LEFT
        if i % 2 == 1:
            ws.cell(row=row, column=1).fill = ZEBRA_FILL
            ws.cell(row=row, column=2).fill = ZEBRA_FILL

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def main() -> None:
    lmin_csv_path = DATA_DIR_SIGMOID / LMIN_MF_FILENAME
    generate_lmin_mf_csv(lmin_csv_path)
    print(f"Generated {lmin_csv_path.name}")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    for spec in VITALS:
        write_sheet(wb, spec)

    write_overview_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
